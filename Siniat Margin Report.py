import pandas as pd
from openpyxl import load_workbook

# Define file path
file_path = r"C:\Users\141823\OneDrive - Etex Group\Desktop\Etex AU - Finance - Finance\06. Revenue & Pricing\Etex\01. Revenue\08. Margin\2025\BW reports\Raw Data.xlsx"

# Load Excel file
xls = pd.ExcelFile(file_path)

# Read "Raw Data" and "Region Mapping" sheets
df = pd.read_excel(xls, sheet_name="Raw Data", header=0)
df_mapping = pd.read_excel(xls, sheet_name="Region Mapping", header=0)

# Merge with Region Mapping to replace "Region" with "BW Region"
df = df.merge(df_mapping, on='Region', how='left')

# Replace "Region" with "BW Region", keeping original values if no match
df['Region'] = df['BW Region'].fillna(df['Region'])

#Drop BW Region now
df = df.drop(columns=['BW Region'])
#-------------------------------------------Product Masterdata----------------------------
import pandas as pd
from openpyxl import load_workbook

# Define file path
file_path = r"C:\Users\141823\OneDrive - Etex Group\Desktop\Etex AU - Finance - Finance\06. Revenue & Pricing\Etex\01. Revenue\08. Margin\2025\BW reports\Raw Data.xlsx"

# Read both material and customer master data
material_df = pd.read_excel(r"C:\Users\141823\Downloads\Material Master Data _Etex AU 5500.xlsx", sheet_name=0, header=None)
customer_df = pd.read_excel(r"C:\Users\141823\Downloads\Customer Master Data _Etex.XLSX", sheet_name='Customer Masterdata')

# Set the second row (index 1) as the column headers
material_df.columns = material_df.iloc[1]  # Set second row as header
material_df = material_df.iloc[2:].reset_index(drop=True)  # Drop first two rows

# Pick needed columns and remove duplicates
material_subset = material_df[['Material', 'Board / Metal','P&L Classification','Group 1 Category','BRAND - Siniat / Promat / Innova / Equitone','Profile Text','Item Group']].rename(
    columns={'BRAND - Siniat / Promat / Innova / Equitone': 'Brand',
             'P&L Classification':'P&L Category',
             'Group 1 Category':'MANUAL CLASSIFICATION (Cornice)'})
# material_subset = material_subset.drop_duplicates(subset=['Material', 'Brand'])

# Merge with Raw Data
df = df.merge(material_subset, on='Material', how='left')

# Filter only for 'SINIAT' and 'GTEK' brands
df = df[df['Brand'].isin(['SINIAT', 'GTEK'])]


#------------------------------------Customer Masterdata-------------------------------
# Select columns by index
customer_subset = customer_df.iloc[:, [0, 9]]

# Rename columns
customer_subset.columns = ['Customer', 'Customer Name Master']
customer_subset=customer_subset.drop_duplicates(subset=['Customer', 'Customer Name Master'])

#Merge with df
df=df.merge(customer_subset, on='Customer', how='left')

# Replace 'Customer Name' in df with 'Customer Name Master' and rename it back to 'Customer Name'
df['Original Customer Name']=df['Customer Name'] 
df['Customer Name'] = df['Customer Name Master']
df['Original Customer']=df['Customer']

# Drop the extra 'Customer Name Master' column
df = df.drop(columns=['Customer Name Master'])


# in Customer Name, Changes every cell wit Keyword 'BUNNINGS' to 'BUNNINGS GROUP LTD' 

df['Customer Name'] = df['Customer Name'].apply(lambda x: 'BUNNINGS GROUP LTD' if 'BUNNINGS' in str(x).upper() else x)
# If 'Customer Name' is 'BUNNINGS GROUP LTD', set 'Customer' to blank
df.loc[df['Customer Name'] == 'BUNNINGS GROUP LTD', 'Customer'] = ""

#add another colum call 'Original Region'

df=df[['Fiscal year/period',	'Customer',	'Customer Name','Material',	'Material Description',	'Region',
      'Plant', 'Actual Cost Of Goods Sold','Quantity','Gross Sales (Invoice)','Transport Surcharge', 'Provisions','Transport to Customers','Interplant Transport',
      'Board / Metal','P&L Category','MANUAL CLASSIFICATION (Cornice)','Profile Text','Item Group',
      'Original Customer','Original Customer Name','Original Region']]

#Save data back to "Raw Data" sheet
# Load the workbook
wb = load_workbook(file_path)
ws = wb["Raw Data"]

# Clear old data but keep headers
ws.delete_rows(2, ws.max_row)  # Deletes from row 2 downwards (keeps headers)

# Write updated DataFrame back to "Raw Data" (including headers)
for c_idx, column in enumerate(df.columns, start=1):
    ws.cell(row=1, column=c_idx, value=column)  # Write column headers

for r_idx, row in enumerate(df.itertuples(index=False), start=2):  # Start from row 2 (below headers)
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column letter (A, B, C, etc.)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

# Save the workbook
wb.save(file_path)

