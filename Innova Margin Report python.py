# BW data frame, Region Mapping and Buying Mapping
import pandas as pd
from openpyxl import load_workbook

# Define file path
file_path = r"C:\Users\141823\OneDrive - Etex Group\Desktop\Etex AU - Finance - Finance\18.Commercial Finance\01. Monthly Reports\Automation-Monthly Margin Report\BW Data.xlsx"

# Load Excel file
xls = pd.ExcelFile(file_path)

# Read "Innova Etex" and "Mapping" sheets
df = pd.read_excel(xls, sheet_name="Innova Etex", header=0)
df_mapping = pd.read_excel(xls, sheet_name="Region Mapping", header=0)
df_buying=pd.read_excel(xls, sheet_name="Buying Mapping", header=0)

# Merge with Region Mapping to replace "Region" with "BW Region"
df = df.merge(df_mapping, on='Region', how='left')
#take useful columns from buying df
df_buying=df_buying[['Customer', 'Buying Group','Buying Group Branch']]
df_buying=df_buying.drop_duplicates(subset=['Customer'])

# Replace "Region" with "BW Region", keeping original values if no match
df['Branch'] = df['Final'].fillna(df['Region'])
#Drop Not needed columns
df = df.drop(columns=['Column V','Final','Note'])

#Merge with buying mapping
df=df.merge(df_buying, on='Customer', how='left')
#-------------------------------------------Product Masterdata----------------------------

# Read both material mater data
material_df = pd.read_excel(r"C:\Users\141823\Downloads\Material Master Data _Etex AU 5500.xlsx", sheet_name=0, header=None)


# Set the second row (index 1) as the column headers
material_df.columns = material_df.iloc[1]  # Set second row as header
material_df = material_df.iloc[2:].reset_index(drop=True)  # Drop first two rows

# Pick needed columns and remove duplicates
material_subset = material_df[['Material', 'P&L Classification','BRAND - Siniat / Promat / Innova / Equitone']].rename(
    columns={'BRAND - Siniat / Promat / Innova / Equitone': 'Brand',
             'P&L Classification':'P&L Category',
             })

material_subset = material_subset.drop_duplicates(subset=['Material', 'Brand'])

# Merge with Raw Data
df = df.merge(material_subset, on='Material', how='left')

# Filter only for 'INNOA' brand
df = df[df['Brand'].isin(['INNOVA'])]


#-------------------item mapping----------------
# Read item mapping into item_df
item_df = pd.read_excel(
    r"C:\Users\141823\OneDrive - Etex Group\Desktop\Etex AU - Finance - Finance\18.Commercial Finance\Reference Files\FC Items Mapping 120225.xlsx",
    header=1,usecols="C,J,K,L")

# Select and rename columns
item_df = item_df[['SAP Item Number.1', 'Product Family Group', 'Product Line Group', 'Product Line Size']].rename(
    columns={
        'SAP Item Number.1': 'Material',
        'Product Family Group': 'Product Family'})

# Drop duplicates based on Material
item_df = item_df.drop_duplicates(subset=['Material'])

# Merge with df
df = df.merge(item_df, on='Material', how='left')


df=df[['Fiscal year/period',	'Customer',	'Customer Name',	'Material',	'Material Description',	
        'Region',	'Plant','Actual Cost Of Goods Sold','Quantity',	'Gross Sales (Invoice)',
        'Transport Surcharge',	'Provisions',	'Transport to Customers',	'Interplant Transport',	
        'P&L Category','Product Family','Product Line Group','Product Line Size','Branch','Brand',
        'Buying Group',	'Buying Group Branch']]

#------------------Prepare Revenue data----------------
revenue_path = r"C:\Users\141823\OneDrive - Etex Group\Desktop\Etex AU - Finance - Finance\18.Commercial Finance\01. Monthly Reports\Automation-Monthly Margin Report\Revenue Monthly 2025.xlsx"

# Load Excel file
xls1 = pd.ExcelFile(revenue_path)

# Read "Revenue sheet and "Mapping" from df_mapping
revenue_df = pd.read_excel(xls1, sheet_name="Revenue Data", header=0, usecols="A:T")
revenue_df=revenue_df.rename(columns={'Region':'Column V'})
df_mapping_new=df_mapping[['Column V','Final']]
revenue_df=revenue_df.merge(df_mapping_new, on='Column V', how='left')
revenue_df['Region']=revenue_df['Final'].fillna(revenue_df['Column V'])

#merge with item_df
revenue_df=revenue_df.merge(item_df, on='Material', how='left')

# Filter only for 'INNOA' brand
revenue_df = revenue_df[revenue_df['Brand'].isin(['INNOVA'])]

#get month
# Make sure 'Billing date' is datetime type
revenue_df['Billing date'] = pd.to_datetime(revenue_df['Billing date'], errors='coerce')

# Extract month abbreviation (e.g., 'Feb') and create new 'Month' column
revenue_df['Month'] = revenue_df['Billing date'].dt.strftime('%b')
revenue_df['Region Final']=revenue_df['Region']

revenue_df=revenue_df[['SOrg.', 'Bill.doc.', 'Billing date', 'Column V',
                       'Sold-to pt', 'Name Sold_to', 'Plnt', 'Material', 
                       'Description', 'Quantity', 'BUn', 'Alt Qty 1', 'Alt UOM 1', 
                       'Revenue ex Freight', 'Revenue inC Freight', 'Freight', 'P&L Classification (14)', 
                       'Brand', 'Region', 'Material Group',  'Product Family', 'Product Line Group', 
                       'Product Line Size', 'Month','Region Final']]

#--------------------Save df to final file------------------
from openpyxl import load_workbook

# Define path
file_path = r"C:\Users\141823\OneDrive - Etex Group\Desktop\Etex AU - Finance - Finance\18.Commercial Finance\01. Monthly Reports\Automation-Monthly Margin Report\BW Data.xlsx"

# Load workbook once
wb = load_workbook(file_path)

# Define a reusable write function
def write_dataframe_to_sheet(wb, sheet_name, df_to_write):
    ws = wb[sheet_name]
    
    # Clear old data except headers
    ws.delete_rows(2, ws.max_row)

    # Write headers
    for c_idx, column in enumerate(df_to_write.columns, start=1):
        ws.cell(row=1, column=c_idx, value=column)

    # Write data rows
    for r_idx, row in enumerate(df_to_write.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

# Call function for each DataFrame
write_dataframe_to_sheet(wb, "Innova Etex", df)
write_dataframe_to_sheet(wb, "Revenue Data", revenue_df)

# Save workbook
wb.save(file_path)
